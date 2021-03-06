import numpy as np
import openpyxl as op
import time
import calcTime as calc
import BrowserController as bc


#シート読み込み
book = op.load_workbook('test.xlsx')

#シート数と名前を取得
print('--------------------------------')
print(len(book.sheetnames))
print('--------------------------------')
for name in book.get_sheet_names():
    print(name)

#セル値の取得
activeSheet = book.active
print(activeSheet.cell(column=1,row=1).value)

CT = calc.CalcTime()

CT.StartTimer()
CT.StopTimer()
CT.Display()


bc = bc.BrowserController()
bc.test()




